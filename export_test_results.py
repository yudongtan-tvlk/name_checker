"""Export per-name checker results for the 'test' tab to an Excel workbook.

Two tabs, identical 4-column schema:
  - "wrong names"   : the checker is fed nameBefore (the misspelled side)
  - "correct names" : the checker is fed nameAfter  (the corrected side)

Columns:
  name before correction      -- the name FED to the checker (before on the
                                 wrong tab, after on the correct tab)
  detected                    -- name-level status string from check():
                                 ok | variant | suspect | misspelled | unparseable
  correction suggestion       -- full name rebuilt with the top-1 suggestion
                                 swapped into each suspect/misspelled token
                                 (blank when no token is flagged)
  original name after correction -- the gold nameAfter

Rows are the spelling-intent pairs from _load_pairs('test') -- exactly the set
the honest evaluation scores. The checker runs at the production operating
point: baked per-country suspect scores + 1b confusion priors + 1c name models
(all mined/built on the 'good' tab, disjoint from 'test').
"""
import sys

import openpyxl

import load_public as L

HEADERS = ["input", "detect result", "country",
           "correction suggestion", "gold corrected name"]
FLAGGED = ("suspect", "misspelled")


def build_checker():
    """Checker at the production operating point, loaded from data/cache/."""
    ck = L.build_checker_from_cache()
    n_conf = (sum(len(tp) for tp in ck.confusions.token_pairs.values())
              if ck.confusions else 0)
    print(f"checker ready from cache: {n_conf} confusion pairs, "
          f"1c models for {len(ck.name_models)} countries", flush=True)
    return ck


def suggestion_for(result):
    """Rebuild the full name with top-1 suggestions swapped into flagged tokens.

    Returns "" when no token was flagged (nothing to correct) -- including the
    1c-only case where the name is structurally 'suspect' but no single token
    carries a suggestion."""
    changed = False
    out = []
    for t in result["tokens"]:
        if t["status"] in FLAGGED and t["suggestions"]:
            out.append(t["suggestions"][0])
            changed = True
        else:
            out.append(t["token"])
    return " ".join(out) if changed else ""


def rows_for(checker, pairs, use_after):
    """One row per pair: feed nameAfter when use_after else nameBefore."""
    rows = []
    for before, after, country in pairs:
        fed = after if use_after else before
        res = checker.check(fed, country)
        rows.append([fed, res["status"], country, suggestion_for(res), after])
    return rows


def _parse_args(argv):
    """-> (tab, out_override). `--tab good|test` (default good); a bare
    positional overrides the auto filename."""
    tab, out = "good", None
    i = 0
    while i < len(argv):
        if argv[i] == "--tab":
            tab = argv[i + 1]
            i += 2
        else:
            out = argv[i]
            i += 1
    return tab, out


def main():
    tab, out_override = _parse_args(sys.argv[1:])
    enabled = list(L.load_config().enabled_countries)
    ctoken = enabled[0] if len(enabled) == 1 else "-".join(enabled)
    out_path = out_override or f"test_results_{ctoken}_{tab}.xlsx"

    # The 'good' tab is the training corpus: nameAfter tokens were merged into
    # the dict (curated) and the pairs trained 1b/1c. Coverage/detect on 'good'
    # are therefore optimistic — use false-positives-on-correct-names + case
    # review as the dev signal; reserve 'test' for unbiased numbers.
    if tab == "good":
        print("NOTE: 'good' is the TRAINING tab — coverage/detect are "
              "optimistic (leakage). Validate with correct-names FPs + case "
              "review; use 'test' for final numbers.", flush=True)

    checker = build_checker()
    pairs, _ = L._load_pairs(tab)
    pairs = [(b, a, c) for b, a, c in pairs if c in set(enabled)]
    print(f"'{tab}' tab spelling pairs ({ctoken}): {len(pairs)}", flush=True)

    wb = openpyxl.Workbook()
    for sheet_name, use_after in (("wrong names", False), ("correct names", True)):
        ws = wb.active if sheet_name == "wrong names" else wb.create_sheet()
        ws.title = sheet_name
        ws.append(HEADERS)
        for row in rows_for(checker, pairs, use_after):
            ws.append(row)
        flagged = sum(1 for r in range(2, ws.max_row + 1)
                      if ws.cell(r, 2).value in FLAGGED)
        print(f"'{sheet_name}': {ws.max_row - 1} rows, {flagged} flagged "
              f"(suspect/misspelled)", flush=True)

    wb.save(out_path)
    print(f"saved -> {out_path}", flush=True)


if __name__ == "__main__":
    sys.exit(main())
