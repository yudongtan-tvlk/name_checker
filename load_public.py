"""TokenSpellChecker + HONEST evaluation harness (reads the cache, builds nothing).

The per-country dictionaries are built ahead of time by `python3 -m namecheck.data`
and live under `data/cache/<country>/` (token<TAB>rank<TAB>is_curated). This module
loads that cache and scores the disjoint 'test' tab. Unlike evaluate.py (which used
the CSV's own corrected names as the dictionary — an upper bound), it only knows
tokens sourced into the cache. It reports:

  1. Coverage    — % of correct (gold) name tokens accepted by the checker.
                   Tokens not accepted would be false-flagged in production.
  2. Detection   — % of misspelled tokens flagged by the checker.
  3. FPR         — name-level false-positive rate (unique corrected names flagged).
  4. Correction  — top-1 suggestion accuracy on misspelled tokens.

Popularity: each token's per-market frequency percentile (derived from the cache
rank, lower rank = more common) is used two ways:
  a. suggestion ranking — combined score = edit_distance + λ·popularity
  b. detection — a very rare in-dict token sitting close to a very common
     name is flagged as "suspect" (e.g. a one-off 'Mohamab' near 'Mohamad')

Sources behind the cache (see DESIGN.md §3): names-dataset, popular_names, JMnedict,
wikidata, CS corrections, tvlk_bookings, kr_given_names.

Usage:
    python3 load_public.py                 # load cache, evaluate 'test'
    python3 load_public.py --serial        # single-process (debug)
    python3 load_public.py --no-confusion  # skip 1b priors
    python3 load_public.py --no-name-model # skip 1c models
    python3 load_public.py --optimize      # per-market SUSPECT_SCORE grid search
"""
import os
import sys
from collections import defaultdict

from namecheck.normalize import fold, tokenize, dedupe_adjacent
from namecheck import classify_intent, split_passengers
from namecheck.romanize import canonical
from namecheck.scorer import norm_score
from namecheck.features import NGramIndex
from namecheck.confusion import ConfusionModel
from namecheck.name_model import NameModel
from namecheck.closed_class import (
    closed_class_validator, korean_acceptor, japanese_native_checker,
)
from namecheck import is_plausible_name

from namecheck.data.config import (
    load_config, COUNTRY_CC, NO_RANK, LAMBDA,
    SUSPECT_SCORE, COMMON_SCORE, SUSPECT_EDIT, SUSPECT_NEAR_EDIT,
    SUSPECT_SCORES_BY_COUNTRY,
    SURNAME_FILES, CONFUSION_SCORE_GAP, ROUTE_NATIVE_MIN,
    EVIDENCE_CUTOFF_BY_COUNTRY, XLSX_PATH, cache_dir, cache_file,
)
from namecheck.data.tokens_io import read_tokens

def load_cache_tokens(countries):
    """Read per-country + regional token caches into checker-shaped dicts.

    Returns (by_country {country: {token: rank}}, regional {token: rank}).
    Reads data/cache/<country>/tokens.txt and data/cache/regional/tokens.txt
    (3-col token<TAB>rank<TAB>is_curated; legacy 2-col tolerated)."""
    by_country = {}
    for c in countries:
        ranks, _ = read_tokens(cache_file(c, "tokens.txt"))
        by_country[c] = ranks
    regional, _ = read_tokens(cache_file("regional", "tokens.txt"))
    return by_country, regional


# --------------------------------------------------------------------------
# Token-level spell checker backed by the public dictionary
# --------------------------------------------------------------------------
def _score_penalty(score):
    """0.0 (score=100, very common) .. 1.0 (score=0, very rare); log scale."""
    # Convert percentile score back to a log-scale penalty so suggestion ranking
    # still reflects the natural power-law distribution of name frequencies.
    # score=100 → frac=1.0 → penalty≈0; score=0 → frac→0 → penalty→1.
    frac = max(score / 100.0, 1e-6)
    import math
    return min(1.0, -math.log10(frac) / 2.0)


def _compute_scores(pool):
    """Return {token: percentile_score 0-100} relative to this pool's distribution.

    Most common token (lowest rank) → 100. Least common → 0. Scores are
    computed per-country pool so Vietnam's 1k tokens and Indonesia's 80k tokens
    are each internally normalized — same threshold works across all markets.
    """
    if not pool:
        return {}
    # NO_RANK tokens pin to 0.0 — always enter the suspect check regardless of
    # market distribution (they have no frequency signal by definition).
    ranked = {t: r for t, r in pool.items() if r < NO_RANK}
    no_rank_tokens = {t for t in pool if pool[t] >= NO_RANK}
    sorted_tokens = sorted(ranked, key=lambda t: ranked[t])
    n = len(sorted_tokens)
    if n == 0:
        return {t: 0.0 for t in pool}
    if n == 1:
        scores = {sorted_tokens[0]: 50.0}
    else:
        scores = {tok: (n - 1 - i) / (n - 1) * 100.0 for i, tok in enumerate(sorted_tokens)}
    for t in no_rank_tokens:
        scores[t] = 0.0
    return scores


class TokenSpellChecker:
    def __init__(self, by_country, regional):
        self.by_country = by_country    # country -> {token: rank}
        self.regional = regional        # {token: rank}
        self._canon = {}                # country -> {canonical: best token}
        self._shared_idx = None         # one index serves all countries
        self._pool = {}                 # country -> {token: rank}
        self._score = {}                # country -> {token: percentile 0-100}
        self.suspect_scores = {}        # country -> SUSPECT_SCORE override
        self.confusions = None          # ConfusionModel (1b) or None
        self.name_models = {}           # country -> NameModel (1c)
        self._aliases = {}              # country -> {typo: canonical_name} (2a)
        self._closed_class = {}         # country -> reject validator or None (2b, VN)
        self._syll_accept = {}          # country -> accept validator or None (KR)
        self._jp_native = {}            # country -> JP mora native-checker (routing)
        self._surnames = {}             # country -> set of valid surnames
        self._surname_canon = {}        # country -> {canonical: best surname}

    def _ensure(self, country):
        if country in self._pool:
            return
        pool = dict(self.regional)
        for t, r in self.by_country.get(country, {}).items():
            pool[t] = min(pool.get(t, NO_RANK), r)
        # 2c: drop low-evidence tokens (rank worse than the per-country cutoff)
        # so unverified long-tail junk is no longer silently accepted.
        cutoff = EVIDENCE_CUTOFF_BY_COUNTRY.get(country)
        if cutoff is not None:
            pool = {t: r for t, r in pool.items() if r < cutoff}
        if self._shared_idx is None:
            self._shared_idx = NGramIndex()
            for t in self.regional:
                self._shared_idx.add(t)
        for t in pool.keys() - self.regional.keys():    # usually empty
            self._shared_idx.add(t)
        canon = {}
        for t, r in pool.items():
            k = canonical(t, country)
            if k not in canon or r < pool[canon[k]]:    # keep most popular
                canon[k] = t
        self._canon[country] = canon
        self._pool[country] = pool
        self._score[country] = _compute_scores(pool)

        # 2a: load typo->name aliases for this country + the regional pool.
        aliases = {}
        for apath in (cache_file(country, "aliases.txt"),
                      cache_file("regional", "aliases.txt")):
            if os.path.exists(apath):
                for line in open(apath):
                    line = line.rstrip("\n")
                    if line:
                        t, _, c = line.partition("\t")
                        if c and t not in pool:   # don't override a valid token
                            aliases.setdefault(t, c)
        self._aliases[country] = aliases

        # 2b: closed-class syllable logic. VN uses single-syllable membership as
        # a REJECT test; KR uses multi-syllable decomposition as an ACCEPT test
        # (a dict-miss token that decomposes into valid syllables is a plausible
        # Korean name, not a misspelling).
        validator = acceptor = jp_native = None
        spath = cache_file(country, "syllables.txt")
        if os.path.exists(spath):
            syllables = {ln.strip() for ln in open(spath) if ln.strip()}
            validator = closed_class_validator(country, syllables)
            if country == "korea":
                acceptor = korean_acceptor(syllables)
            elif country == "japan":
                jp_native = japanese_native_checker(syllables)
        self._closed_class[country] = validator
        self._syll_accept[country] = acceptor
        self._jp_native[country] = jp_native

        surnames = set()
        fn = SURNAME_FILES.get(country)
        if fn:
            spath = cache_file(country, fn)
            if os.path.exists(spath):
                surnames = {ln.strip().lower() for ln in open(spath) if ln.strip()}
        scanon = {}
        scores = self._score[country]
        for s in surnames:
            k = canonical(s, country)
            cur = scanon.get(k)
            if cur is None or scores.get(s, 0.0) > scores.get(cur, 0.0):
                scanon[k] = s
        self._surnames[country] = surnames
        self._surname_canon[country] = scanon

    def _suggest(self, tok, country, top_k):
        """Popularity-weighted nearest names: edit + λ·score_penalty."""
        ctok = canonical(tok, country)
        scores = self._score[country]
        scored = []
        for c in self._shared_idx.candidates(tok, top_k=60):
            s = scores.get(c)
            if s is None:      # token belongs to another country's dict only
                continue
            edit = norm_score(ctok, canonical(c, country))
            scored.append((edit + LAMBDA * _score_penalty(s), edit, c))
        scored.sort()
        return scored[:top_k]

    def slots(self, toks, country):
        """Positional slots (1a). The surname slot is token 0 of a 2+ token
        name — but only when no LATER token matches the surname inventory.
        If one does, the name is probably not surname-first order (common in
        KR bookings: 'jihun lee'), so the slot grammar is skipped."""
        self._ensure(country)
        surnames = self._surnames.get(country)
        slots = [None] * len(toks)
        if not surnames or len(toks) < 2:
            return slots
        # A single-letter first token is an initial, not a surname-grammar
        # violation -> don't impose the surname slot on it.
        if len(toks[0]) < 2:
            return slots
        scanon = self._surname_canon[country]
        if not any(t in surnames or canonical(t, country) in scanon
                   for t in toks[1:]):
            slots[0] = "surname"
        return slots

    def _confusion_neighbor(self, tok, country):
        """1b char-level: a strictly-more-common in-dict neighbor reachable
        from tok via frequently-mined confusion ops, or None."""
        scores = self._score[country]
        t_score = scores.get(tok, 0.0)
        # Partner must be clearly more common, not just marginally — a soft
        # confirm on a near-tie would flag half the valid name space.
        best, best_score = None, t_score + CONFUSION_SCORE_GAP
        for c in self._shared_idx.candidates(tok, top_k=30):
            if c == tok:
                continue
            s = scores.get(c)
            if s is None or s < best_score:
                continue
            if self.confusions.ops_confusable(tok, c, country):
                best, best_score = c, s
        return best

    def _suggest_surnames(self, tok, country, top_k):
        """Nearest surnames only — the surname slot draws from a closed list."""
        ctok = canonical(tok, country)
        scores = self._score[country]
        scored = []
        for s in self._surnames[country]:
            edit = norm_score(ctok, canonical(s, country))
            scored.append((edit + LAMBDA * _score_penalty(scores.get(s, 0.0)), edit, s))
        scored.sort()
        return [s for _, _, s in scored[:top_k]]

    def check_token(self, tok, country, top_k=3, slot=None):
        """-> (status, suggestions). status: ok | variant | suspect | misspelled

        slot="surname" applies the positional slot grammar (1a): for countries
        with a closed surname inventory, the token must be a valid surname;
        a valid *given* name in the surname slot is 'suspect', an unknown
        token is 'misspelled', and suggestions come from the surname list.
        """
        self._ensure(country)
        # 2a: a quarantined typo is a known misspelling of a real name.
        alias = self._aliases.get(country, {}).get(tok)
        if alias is not None:
            return "misspelled", [alias]
        # 2b: a token that is not a valid syllable of a closed-class language
        # (e.g. Vietnamese 'thithuy', 'kk') cannot be a real name token, even
        # if a public dump enshrined it. Flag regardless of dict membership.
        validator = self._closed_class.get(country)
        if validator is not None and len(tok) >= 2 and tok.isalpha() and not validator(tok):
            sugg = [c for _, _, c in self._suggest(tok, country, top_k)]
            status = "suspect" if tok in self._score[country] else "misspelled"
            return status, sugg
        if slot == "surname" and self._surnames.get(country):
            if tok in self._surnames[country]:
                return "ok", []
            ckey = canonical(tok, country)
            if ckey in self._surname_canon[country]:
                return "variant", [self._surname_canon[country][ckey]]
            sugg = self._suggest_surnames(tok, country, top_k)
            status = "suspect" if tok in self._score[country] else "misspelled"
            return status, sugg
        scores = self._score[country]
        score = scores.get(tok)
        if score is not None:
            # 1b: in-dict token on the high-risk side of a mined confusion
            # -> soft 'suspect' with the confusion partner as suggestion.
            # Guard: never downgrade a token that is itself COMMON. CS logs
            # correct names in both directions, so token_hit can otherwise flag
            # the most common name in a market (e.g. VN 'anh'/'tran') just
            # because the reverse correction was also logged. A high-percentile
            # token is the canonical form, not a typo of a rarer one.
            if self.confusions is not None and score < COMMON_SCORE:
                g = self.confusions.token_hit(tok, country)
                if g is None:
                    g = self._confusion_neighbor(tok, country)
                if g:
                    return "suspect", [g]
            thresh = self.suspect_scores.get(country, SUSPECT_SCORE)
            if score >= thresh:
                return "ok", []
            # bottom-percentile token: suspect if a high-percentile name is very
            # close. Skipped for 'regional' (foreign-routed) names -- the weak
            # cross-Asian pool is no basis to challenge a foreign name.
            if country != "regional":
                for _, edit, c in self._suggest(tok, country, top_k):
                    if c != tok and edit <= SUSPECT_NEAR_EDIT and scores.get(c, 0.0) >= COMMON_SCORE:
                        return "suspect", [c]
            return "ok", []
        ckey = canonical(tok, country)
        if ckey in self._canon[country]:
            return "variant", [self._canon[country][ckey]]
        # A: dict-miss but decomposes into valid Korean syllables -> a plausible
        # romanized KR name we simply don't have listed; accept rather than flag
        # a fabricated suggestion. BUT only when no close, common dictionary name
        # exists: a near-and-common anchor means this is more likely a typo of
        # that name than a novel one, so keep it flagged (with the anchor as the
        # suggestion). This recovers detection of typos-of-common-names that the
        # raw syllable test would otherwise swallow.
        sugg = self._suggest(tok, country, top_k)
        acc = self._syll_accept.get(country)
        if acc is not None and acc(tok):
            near = next((c for _, edit, c in sugg
                         if c != tok and edit <= SUSPECT_EDIT
                         and scores.get(c, 0.0) >= COMMON_SCORE), None)
            if near is None:
                return "ok", []
            return "misspelled", [near]
        # foreign-routed name: a dict-miss against the weak regional pool is not
        # evidence of a typo, so don't hard-flag it.
        if country == "regional":
            return "ok", []
        return "misspelled", [c for _, _, c in sugg]

    def route(self, toks, country):
        """3c: pick the rule-set by NAME, not booking origin. origin_country
        is the market, not the passenger's nationality. A foreign name on a
        VN/KR booking should not be judged by VN/KR closed-inventory rules.
        Returns the effective country ('regional' = neutral, no closed-class
        / surname-slot) for checking; reporting still uses the booking origin.
        """
        country = (country or "").lower()
        alpha = [t for t in toks if len(t) >= 2 and t.isalpha()]
        if not alpha:
            return country
        self._ensure(country)
        if country == "vietnam":
            val = self._closed_class.get("vietnam")
            if val is not None:
                native = sum(1 for t in alpha if val(t)) / len(alpha)
                if native < ROUTE_NATIVE_MIN:
                    return "regional"
        elif country == "korea":
            own = self.by_country.get("korea", {})
            scanon = self._surname_canon.get("korea", {})
            has_surname = any(canonical(t, "korea") in scanon for t in alpha)
            in_own = sum(1 for t in alpha if t in own) / len(alpha)
            if not has_surname and in_own < ROUTE_NATIVE_MIN:
                return "regional"
        elif country == "japan":
            # The JP dict (jmnedict ~242k) "contains" almost everything, so
            # membership can't detect foreign names. Use Japanese-mora
            # phonotactics instead: a name with too few native-romaji tokens
            # is a foreign passenger -> neutral 'regional' rules.
            nat = self._jp_native.get("japan")
            if nat is not None:
                native = sum(1 for t in alpha if nat(t)) / len(alpha)
                if native < ROUTE_NATIVE_MIN:
                    return "regional"
        return country

    def name_anomalous(self, toks, country):
        """1c: True if the per-country name model flags this token sequence.

        A bare 'duplicate' flag on a COMMON name component is not an anomaly:
        reduplication of common names is a real pattern (Sikh 'Singh ... Singh',
        VN 'Nguyen Khoi Nguyen', Burmese 'Si Si'). Only a duplicated *rare*
        token (likely a doubled-surname data error) keeps the flag.
        """
        m = self.name_models.get(country)
        if not m:
            return False
        self._ensure(country)
        scores = self._score.get(country, {})
        for typ, detail in m.anomalies(toks):
            if typ == "duplicate":
                dups = detail.split(",")
                if all(scores.get(d, 0.0) >= COMMON_SCORE for d in dups):
                    continue   # common reduplication -> not anomalous
            return True
        return False

    def segment(self, toks, country):
        """Split a single space-less name into surname + given using the closed
        surname inventory as a prefix/suffix anchor (e.g. 'limkunhee' -> 'lim',
        'kunhee'; 'nahyunkim' -> 'nahyun', 'kim').

        Only fires on a SINGLE-token input (the no-space concatenation case) so
        it never re-splits a token inside an already-spaced name. Guardrails:
        the token must be long (>=7), not itself a valid dict token or surname,
        and each side must be >=3 chars. The longest surname match wins; a
        prefix match is preferred over a suffix one on ties."""
        if len(toks) != 1:
            return toks
        self._ensure(country)
        surnames = self._surnames.get(country)
        if not surnames:
            return toks
        t = toks[0]
        if len(t) < 7 or t in self._pool.get(country, {}) or t in surnames:
            return toks
        # rank candidates by (surname length, prefix-preferred) so ties are
        # deterministic and a leading surname wins over a trailing one.
        best_key, best = None, None
        for s in surnames:
            if len(s) < 2 or len(t) - len(s) < 3:
                continue
            if t.startswith(s):
                key = (len(s), 1)
                if best_key is None or key > best_key:
                    best_key, best = key, [s, t[len(s):]]
            if t.endswith(s):
                key = (len(s), 0)
                if best_key is None or key > best_key:
                    best_key, best = key, [t[:-len(s)], s]
        return best if best else toks

    def tokens(self, raw, country):
        """tokenize + country-aware surname segmentation (the canonical token
        list every checking path should use, so coverage/detect/export agree)."""
        return self.segment(tokenize(raw), country)

    def check(self, raw, country):
        if not is_plausible_name(raw):
            return {"status": "unparseable", "tokens": []}
        toks = self.tokens(raw, country)
        eff = self.route(toks, country)        # 3c: rule-set by name
        slots = self.slots(toks, eff)
        out = [(t, *self.check_token(t, eff, slot=sl))
               for t, sl in zip(toks, slots)]
        order = {"misspelled": 3, "suspect": 2, "variant": 1, "ok": 0}
        worst = max((s for _, s, _ in out), key=order.get, default="ok")
        # 1c: structural anomaly escalates an otherwise-clean name to 'suspect'.
        if order.get(worst, 0) < order["suspect"] and self.name_anomalous(toks, eff):
            worst = "suspect"
        return {"status": worst,
                "tokens": [{"token": t, "status": s, "suggestions": sg}
                           for t, s, sg in out]}


# --------------------------------------------------------------------------
# Honest evaluation against the BQ CSV
# --------------------------------------------------------------------------
def _aligned_token_pairs(before, after, country, checker=None):
    """Greedy-align tokens of a (misspelled, correct) pair; return
    (wrong, gold, slot) triples that differ."""
    # 3b: duplicated-token collapse; segment space-less names when a checker is
    # available (so 'limkunhee' aligns as lim+kunhee).
    _tok = (lambda s: dedupe_adjacent(checker.tokens(s, country))) if checker \
        else (lambda s: dedupe_adjacent(tokenize(s)))
    tb = _tok(before)
    ta = _tok(after)
    slots = checker.slots(tb, country) if checker else [None] * len(tb)
    pairs, used = [], set()
    for i, w in enumerate(tb):
        best, bj = None, None
        for j, g in enumerate(ta):
            if j in used:
                continue
            s = norm_score(canonical(w, country), canonical(g, country))
            if best is None or s < best:
                best, bj = s, j
        if bj is not None and best is not None and best < 0.5:
            used.add(bj)
            if w != ta[bj]:
                pairs.append((w, ta[bj], slots[i]))
    return pairs


def _process_rows(rows):
    """Per-row routing (the CPU-heavy part: classify_intent runs edit distance).

    Pure function over a list of raw (before, after, country) triples -> the
    surviving spelling pairs plus a routed-count dict. Order-preserving and
    deterministic, so chunking it across processes is bit-identical to serial."""
    pairs, routed = [], defaultdict(int)
    for b, a, c in rows:
        bs = split_passengers(b) or [b]
        as_ = split_passengers(a) or [a]
        if len(bs) != len(as_):
            routed["unaligned_multi"] += 1
            continue
        for bb, aa in zip(bs, as_):
            bb, aa = bb.strip(), aa.strip()
            if not (is_plausible_name(bb) and is_plausible_name(aa)):
                continue
            intent = classify_intent(bb, aa, c)
            if intent != "spelling":
                routed[intent] += 1
                continue
            pairs.append((bb, aa, c))
    return pairs, dict(routed)


def _load_pairs(sheet="good"):
    """Load (before, after, country) spelling pairs from a tab of XLSX_PATH.

    Multi-passenger rows (literal \\n separators) are split per passenger
    and aligned pairwise; rows whose sides have different passenger counts
    are diverted as unalignable. The per-row routing (classify_intent, which
    is edit-distance heavy) dominates and is fanned out over process chunks;
    the xlsx read itself is cheap and stays serial."""
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    ws = wb[sheet]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    idx = {h: i for i, h in enumerate(header)}
    raw = []
    for r in rows_iter:
        b = str(r[idx["nameBefore"]] or "").strip()
        a = str(r[idx["nameAfter"]] or "").strip()
        c = str(r[idx["origin_country"]] or "").strip().lower()
        if c not in COUNTRY_CC:
            continue
        raw.append((b, a, c))
    wb.close()

    n_workers = 1 if "--serial" in sys.argv else min(os.cpu_count() or 1, 8)
    if n_workers <= 1 or len(raw) < 256:
        pairs, routed = _process_rows(raw)
        return pairs, defaultdict(int, routed)

    # Even chunks, processed and concatenated in order -> identical to serial.
    chunks, step = [], -(-len(raw) // n_workers)   # ceil division
    for i in range(0, len(raw), step):
        chunks.append(raw[i:i + step])
    pairs, routed = [], defaultdict(int)
    try:
        from concurrent.futures import ProcessPoolExecutor
        with ProcessPoolExecutor(max_workers=n_workers) as ex:
            results = list(ex.map(_process_rows, chunks))
    except Exception as e:                       # pragma: no cover
        print(f"[parallel] _load_pairs pool failed ({e}); serial", flush=True)
        results = [_process_rows(c) for c in chunks]
    for cp, cr in results:
        pairs.extend(cp)
        for k, v in cr.items():
            routed[k] += v
    return pairs, routed


def _eval_country_metrics(checker, cpairs, country):
    """Return (n_gold, covered, n_wrong, detected, n_wrong_in, corr_in, corr_all,
    n_names, n_fp). The last two drive name-level false-positive rate: a unique
    corrected name is a false positive if it is structurally flagged (1c) or any
    of its tokens comes back 'suspect'/'misspelled' (symmetric with the detect
    side). Deduped per country so common names aren't over-counted; coverage and
    detect remain per-occurrence."""
    n_gold = covered = 0
    seen_names = set()
    n_names = n_fp = 0
    for _, a in cpairs:
        ta = checker.tokens(a, country)         # tokenize + surname segmentation
        eff = checker.route(ta, country)        # 3c: rule-set by name
        slots = checker.slots(ta, eff)
        # 1c: a structurally-flagged gold name goes to review -> its tokens
        # are not silently accepted (counts against coverage).
        name_anom = checker.name_anomalous(ta, eff)
        name_ok = not name_anom
        key = tuple(ta)
        first_seen = key not in seen_names
        if first_seen:
            seen_names.add(key)
            n_names += 1
        name_fp = name_anom
        for t, sl in zip(ta, slots):
            n_gold += 1
            status = checker.check_token(t, eff, slot=sl)[0]
            if name_ok and status in ("ok", "variant"):
                covered += 1
            if status in ("suspect", "misspelled"):
                name_fp = True
        if first_seen and name_fp:
            n_fp += 1
    detected = corr_in = corr_all = n_wrong = n_wrong_in = 0
    for b, a in cpairs:
        tb = checker.tokens(b, country)
        eff = checker.route(tb, country)
        # 1c: structural flag on the misspelled name detects every error in it.
        name_flag = checker.name_anomalous(tb, eff)
        for wrong, gold, slot in _aligned_token_pairs(b, a, eff, checker):
            n_wrong += 1
            status, sugg = checker.check_token(wrong, eff, slot=slot)
            hit = bool(sugg) and fold(sugg[0]) == fold(gold)
            if name_flag or status in ("misspelled", "variant", "suspect"):
                detected += 1
            if checker.check_token(gold, eff, slot=slot)[0] in ("ok", "variant"):
                n_wrong_in += 1
                corr_in += hit
            corr_all += hit
    return (n_gold, covered, n_wrong, detected, n_wrong_in, corr_in, corr_all,
            n_names, n_fp)


_SWEEP_VALUES = [1, 2, 3, 5, 7, 10, 15, 20, 25, 30, 40, 50]


OPT_SAMPLE_CAP = 1500   # per-country pair cap for the sweep; evenly-spaced
                        # deterministic sample (full-data sweep takes hours)


# --------------------------------------------------------------------------
# Parallel country dispatch
#
# The eval/optimize country loops are embarrassingly parallel: each country
# owns a private dictionary, NGram index, and metric accumulation. We run one
# process per country. Workers rebuild a single-country checker from the disk
# cache (cheap) rather than pickling the multi-country indexes; the small 1b/1c
# model objects ARE passed through (they pickle to a few KB). Each country in
# isolation yields bit-identical numbers because _suggest already filters
# foreign-dictionary candidates via the per-country score lookup.
# --------------------------------------------------------------------------
_WORKER_STATE = {}


def _worker_init():
    """Per-process one-time load of the cached dictionaries (stdout suppressed)."""
    import io, contextlib
    with contextlib.redirect_stdout(io.StringIO()):
        cfg = load_config()
        by_country, regional = load_cache_tokens(cfg.enabled_countries)
    _WORKER_STATE["by_country"] = by_country
    _WORKER_STATE["regional"] = regional


def _make_country_checker(country, suspect_score, confusion_model, name_model):
    if not _WORKER_STATE:
        _worker_init()
    st = _WORKER_STATE
    ck = TokenSpellChecker({country: st["by_country"].get(country, {})},
                           st["regional"])
    ck.suspect_scores = {country: suspect_score}
    ck.confusions = confusion_model
    ck.name_models = {country: name_model} if name_model else {}
    ck._ensure(country)
    return ck


def _worker_eval(payload):
    country, cpairs, suspect_score, conf, nm = payload
    ck = _make_country_checker(country, suspect_score, conf, nm)
    return country, _eval_country_metrics(ck, cpairs, country)


def _worker_sweep(payload):
    country, cpairs, sweep_values, conf, nm = payload
    ck = _make_country_checker(country, sweep_values[0], conf, nm)
    best_obj, best_val, rows = -1, sweep_values[0], []
    for val in sweep_values:
        ck.suspect_scores[country] = val
        ng, cov, nw, det, nwi, ci, ca, nn, nfp = _eval_country_metrics(ck, cpairs, country)
        det_r = det / nw if nw else 0
        ca_r = ca / nw if nw else 0
        fpr = nfp / nn if nn else 0
        obj = det_r + ca_r - fpr
        rows.append((val, det_r, ca_r, fpr, obj))
        if obj > best_obj:
            best_obj, best_val = obj, val
    return country, rows, best_val


def _dispatch(payloads, worker):
    """Map worker over per-country payloads, parallel unless --serial / trivial.

    Falls back to serial on any pool error (e.g. an un-picklable payload) so a
    parallelism problem never blocks the evaluation."""
    if "--serial" in sys.argv or len(payloads) <= 1:
        return [worker(p) for p in payloads]
    from concurrent.futures import ProcessPoolExecutor
    n = min(os.cpu_count() or 1, len(payloads))
    try:
        with ProcessPoolExecutor(max_workers=n, initializer=_worker_init) as ex:
            return list(ex.map(worker, payloads))
    except Exception as e:                       # pragma: no cover
        print(f"[parallel] pool failed ({e}); running serially", flush=True)
        return [worker(p) for p in payloads]


def optimize_suspect_scores(checker, pairs):
    """Grid-search SUSPECT_SCORE per country. Objective: detect + corr@1-all - fpr
    (name-level false-positive rate penalized equally)."""
    print("\n=== OPTIMIZING per-country SUSPECT_SCORE ===")
    print(f"sweep values: {_SWEEP_VALUES}")
    print(f"objective: detect + corr@1-all - fpr (higher = better)\n", flush=True)
    payloads = []
    for country in COUNTRY_CC:
        cpairs = [(b, a) for b, a, c in pairs if c == country]
        if not cpairs:
            continue
        n_orig = len(cpairs)
        if n_orig > OPT_SAMPLE_CAP:
            step = n_orig / OPT_SAMPLE_CAP
            cpairs = [cpairs[int(i * step)] for i in range(OPT_SAMPLE_CAP)]
            print(f"  {country}: sampled {OPT_SAMPLE_CAP} of {n_orig} pairs", flush=True)
        nm = checker.name_models.get(country)
        payloads.append((country, cpairs, _SWEEP_VALUES, checker.confusions, nm))
    results = {c: (rows, best) for c, rows, best in _dispatch(payloads, _worker_sweep)}
    optimal = {}
    for country in COUNTRY_CC:                  # print in stable order
        if country not in results:
            continue
        rows, best_val = results[country]
        checker.suspect_scores[country] = best_val
        optimal[country] = best_val
        print(f"  {country}:")
        for val, det_r, ca_r, fpr, obj in rows:
            marker = " <-- optimal" if val == best_val else ""
            print(f"    score={val:>3}  detect={det_r:5.1%}  corr@1-all={ca_r:5.1%}  "
                  f"fpr={fpr:5.1%}  obj={obj:.4f}{marker}")
        sys.stdout.flush()
    print(f"\noptimal SUSPECT_SCORE per country: {optimal}\n")
    return optimal


def _load_gold_names(sheet="good"):
    """-> {country: [token_list, ...]} from nameAfter of the given tab."""
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    idx = {h: i for i, h in enumerate(header)}
    out = defaultdict(list)
    for r in it:
        a = str(r[idx["nameAfter"]] or "").strip()
        c = str(r[idx["origin_country"]] or "").strip().lower()
        if not a or c not in COUNTRY_CC:
            continue
        for name in (split_passengers(a) or [a]):
            toks = tokenize(name)
            if toks:
                out[c].append(toks)
    wb.close()
    return out


def build_name_models(gold_by_country):
    """1c: one positional-unigram+bigram NameModel per country."""
    models = {}
    for country, names in gold_by_country.items():
        if len(names) >= 50:        # too few names => no usable distribution
            models[country] = NameModel().fit(names)
    return models


def _name_struct_eval(models, test_pairs):
    """Honest name-level structural metric (1c). For each spelling pair:
    is the misspelled-side name flagged anomalous (struct-detect) and is the
    gold-side name wrongly flagged (struct-FP)? Models trained on 'good',
    pairs from 'test' — disjoint, so no leakage."""
    print("\n=== 1c name-level structural model (trained on 'good', test on 'test') ===")
    hdr = f"{'country':12} {'pairs':>6} {'struct-detect':>14} {'struct-FP':>11} {'model-N':>8}"
    print(hdr + "\n" + "-" * len(hdr))
    tot_d = tot_fp = tot_n = 0
    for country in COUNTRY_CC:
        m = models.get(country)
        cpairs = [(b, a) for b, a, c in test_pairs if c == country]
        if not m or not cpairs:
            continue
        det = sum(1 for b, _ in cpairs if m.is_anomalous(tokenize(b)))
        fp = sum(1 for _, a in cpairs if m.is_anomalous(tokenize(a)))
        print(f"{country:12} {len(cpairs):>6} {det/len(cpairs):>13.1%} "
              f"{fp/len(cpairs):>10.1%} {m.n_train:>8}")
        tot_d += det; tot_fp += fp; tot_n += len(cpairs)
    print("-" * len(hdr))
    if tot_n:
        print(f"{'TOTAL':12} {tot_n:>6} {tot_d/tot_n:>13.1%} {tot_fp/tot_n:>10.1%}")
    print("""
struct-detect : misspelled-side names flagged as structurally anomalous
struct-FP     : gold-side names wrongly flagged (the friction cost)
""")


def _mine_confusions(train_pairs, checker):
    """Aggregate (wrong, gold, country) token triples into a ConfusionModel."""
    triples = []
    for b, a, c in train_pairs:
        for wrong, gold, _slot in _aligned_token_pairs(b, a, c, checker):
            triples.append((wrong, gold, c))
    return ConfusionModel().mine(triples)


def evaluate(checker, pairs=None, routed=None, confusion_cv=0):
    """confusion_cv >= 2 enables 1b priors, k-fold cross-validated: each
    fold is scored with a model mined only on the other folds (the eval CSV
    is also the mining corpus — without CV the numbers would be leakage)."""
    if pairs is None:
        pairs, routed = _load_pairs()
        print(f"\n3a routing: {sum(routed.values())} pairs diverted ({dict(routed)})")
    models = None
    nf = confusion_cv
    if confusion_cv >= 2:
        models = [
            _mine_confusions([p for i, p in enumerate(pairs) if i % nf != f], checker)
            for f in range(nf)
        ]
        print(f"\n1b confusion priors active ({nf}-fold cross-validated mining)")
    print(f"\n=== HONEST evaluation: public dictionary only, {len(pairs)} pairs ===")
    hdr = (f"{'country':12} {'gold-tok':>8} {'coverage':>9} {'detect':>7} "
           f"{'fpr':>6} {'corr@1*':>8} {'corr@1-all':>10} {'sus_score':>10}")
    print(hdr + "\n" + "-" * len(hdr))
    # Non-CV path: one process per country (the CV path stays serial because
    # each fold re-binds checker.confusions and is rarely used).
    par_metrics = {}
    if not models:
        payloads = []
        for country in COUNTRY_CC:
            cpairs = [(b, a) for b, a, c in pairs if c == country]
            if not cpairs:
                continue
            ss = checker.suspect_scores.get(country, SUSPECT_SCORE)
            nm = checker.name_models.get(country)
            payloads.append((country, cpairs, ss, checker.confusions, nm))
        par_metrics = {c: m for c, m in _dispatch(payloads, _worker_eval)}
    tot = defaultdict(float)
    for country in COUNTRY_CC:
        if models:
            m9 = [0] * 9
            for f in range(nf):
                fold_cpairs = [(b, a) for i, (b, a, c) in enumerate(pairs)
                               if c == country and i % nf == f]
                checker.confusions = models[f]
                m = _eval_country_metrics(checker, fold_cpairs, country)
                m9 = [x + y for x, y in zip(m9, m)]
            checker.confusions = None
            ng, cov, nw, det, nwi, ci, ca, nn, nfp = m9
        else:
            if country not in par_metrics:
                continue
            ng, cov, nw, det, nwi, ci, ca, nn, nfp = par_metrics[country]
        if not ng:
            continue
        cov_r = cov / ng if ng else 0
        det_r = det / nw if nw else 0
        fpr   = nfp / nn if nn else 0
        c_in  = ci / nwi if nwi else 0
        c_all = ca / nw if nw else 0
        ss = checker.suspect_scores.get(country, SUSPECT_SCORE)
        print(f"{country:12} {ng:>8} {cov_r:>8.1%} {det_r:>6.1%} "
              f"{fpr:>5.1%} {c_in:>7.1%} {c_all:>9.1%} {ss:>10.1f}")
        for k, v in (("gt", ng), ("cov", cov), ("nw", nw),
                     ("det", det), ("nwi", nwi), ("ci", ci), ("ca", ca),
                     ("nn", nn), ("nfp", nfp)):
            tot[k] += v
    print("-" * len(hdr))
    print(f"{'TOTAL':12} {int(tot['gt']):>8} {tot['cov']/tot['gt']:>8.1%} "
          f"{tot['det']/tot['nw']:>6.1%} {tot['nfp']/tot['nn']:>5.1%} "
          f"{tot['ci']/tot['nwi']:>7.1%} {tot['ca']/tot['nw']:>9.1%}")
    print("""
coverage   : correct tokens accepted (1-coverage = token-level false-flag rate)
detect     : misspelled tokens flagged (incl. rare-but-in-dict 'suspect')
fpr        : name-level false-positive rate -- % of UNIQUE corrected names
             (nameAfter) flagged, i.e. structurally anomalous (1c) or any token
             'suspect'/'misspelled'. Deduped per country; coverage/detect are
             per-occurrence, so fpr is not simply 1-coverage.
corr@1*    : top-1 correction accuracy WHEN the gold token is in the dict
corr@1-all : top-1 correction accuracy over ALL misspelled tokens
""")


def build_checker_from_cache(config=None, with_confusions=True, with_name_models=True):
    """Construct a TokenSpellChecker from the per-country cache (data/cache/).

    Loads tokens for the config's enabled countries (+ regional fallback), the
    baked per-market suspect scores, and — unless disabled — the cached 1b
    confusion model and 1c name models. No mining/building happens here; the
    artifacts are produced by `python3 -m namecheck.data`.
    """
    config = config or load_config()
    countries = config.enabled_countries
    by_country, regional = load_cache_tokens(countries)
    checker = TokenSpellChecker(by_country, regional)
    checker.suspect_scores = dict(SUSPECT_SCORES_BY_COUNTRY)
    if with_confusions:
        cpaths = {c: (cache_file(c, "confusions.tsv"),
                      cache_file(c, "confusion_ops.tsv")) for c in countries}
        checker.confusions = ConfusionModel.load(cpaths)
    if with_name_models:
        nm = {}
        for c in countries:
            p = cache_file(c, "namemodel.json")
            if os.path.exists(p):
                nm[c] = NameModel.load(p)
        checker.name_models = nm
    return checker


if __name__ == "__main__":
    config = load_config()
    with_conf = "--no-confusion" not in sys.argv
    with_nm = "--no-name-model" not in sys.argv
    checker = build_checker_from_cache(config, with_confusions=with_conf,
                                       with_name_models=with_nm)
    sizes = {c: len(checker.by_country.get(c, {})) for c in config.enabled_countries}
    print(f"enabled countries: {config.enabled_countries}")
    print(f"dictionary sizes: {sizes}\nregional pool: {len(checker.regional)} tokens")
    if checker.confusions is not None:
        n_tok = sum(len(tp) for tp in checker.confusions.token_pairs.values())
        print(f"1b confusion model loaded from cache: {n_tok} token confusions")
    if checker.name_models:
        n_built = ", ".join(f"{c}:{m.n_train}" for c, m in sorted(checker.name_models.items()))
        print(f"1c name models loaded from cache: [{n_built}]")

    test_pairs, test_routed = _load_pairs("test")
    test_pairs = [(b, a, c) for b, a, c in test_pairs
                  if c in config.enabled_countries]
    print(f"\ntest tab: {len(test_pairs)} spelling pairs (enabled countries), "
          f"{sum(test_routed.values())} diverted ({dict(test_routed)})")
    if "--optimize" in sys.argv:
        good_pairs, _ = _load_pairs("good")
        good_pairs = [p for p in good_pairs if p[2] in config.enabled_countries]
        optimize_suspect_scores(checker, good_pairs)
    evaluate(checker, test_pairs, test_routed)
    if checker.name_models:
        _name_struct_eval(checker.name_models, test_pairs)
